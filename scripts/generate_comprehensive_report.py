#!/usr/bin/env python3
"""
Comprehensive Extraction Report Generator
Generates detailed Excel reports with multiple sheets and statistics
"""
import sys
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
from collections import defaultdict

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from intelligence_capture.database import IntelligenceDB
from intelligence_capture.config import DB_PATH


class ExtractionReportGenerator:
    """Generates comprehensive extraction reports"""

    def __init__(self, db_path: Path = DB_PATH):
        self.db = IntelligenceDB(db_path)
        self.db.connect()
        self.report_data = {}

    def generate_full_report(self) -> Dict[str, Any]:
        """Generate comprehensive report with all statistics"""
        print(f"\n{'='*70}")
        print(f"📊 GENERATING COMPREHENSIVE EXTRACTION REPORT")
        print(f"{'='*70}\n")

        # Collect all statistics
        self.report_data = {
            "metadata": self._get_metadata(),
            "summary": self._get_summary_stats(),
            "entity_counts": self._get_entity_counts(),
            "company_breakdown": self._get_company_breakdown(),
            "quality_metrics": self._get_quality_metrics(),
            "top_entities": self._get_top_entities(),
            "extraction_status": self._get_extraction_status()
        }

        return self.report_data

    def _get_metadata(self) -> Dict[str, Any]:
        """Get report metadata"""
        return {
            "generated_at": datetime.now().isoformat(),
            "database": str(self.db.db_path),
            "report_version": "1.0"
        }

    def _get_summary_stats(self) -> Dict[str, Any]:
        """Get high-level summary statistics"""
        cursor = self.db.conn.cursor()

        # Count interviews
        cursor.execute("SELECT COUNT(*) FROM interviews")
        total_interviews = cursor.fetchone()[0]

        # Count companies
        cursor.execute("SELECT COUNT(DISTINCT company) FROM interviews")
        total_companies = cursor.fetchone()[0]

        # Count total entities across all types
        entity_tables = [
            "pain_points", "processes", "systems", "kpis",
            "automation_candidates", "inefficiencies",
            "communication_channels", "decision_points", "data_flows",
            "temporal_patterns", "failure_modes", "team_structures",
            "knowledge_gaps", "success_patterns", "budget_constraints",
            "external_dependencies"
        ]

        total_entities = 0
        for table in entity_tables:
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                total_entities += cursor.fetchone()[0]
            except:
                pass

        return {
            "total_interviews": total_interviews,
            "total_companies": total_companies,
            "total_entities": total_entities,
            "avg_entities_per_interview": total_entities / total_interviews if total_interviews > 0 else 0
        }

    def _get_entity_counts(self) -> Dict[str, int]:
        """Get entity counts for all types"""
        cursor = self.db.conn.cursor()

        entity_tables = {
            "pain_points": "Pain Points",
            "processes": "Processes",
            "systems": "Systems",
            "kpis": "KPIs",
            "automation_candidates": "Automation Candidates",
            "inefficiencies": "Inefficiencies",
            "communication_channels": "Communication Channels",
            "decision_points": "Decision Points",
            "data_flows": "Data Flows",
            "temporal_patterns": "Temporal Patterns",
            "failure_modes": "Failure Modes",
            "team_structures": "Team Structures",
            "knowledge_gaps": "Knowledge Gaps",
            "success_patterns": "Success Patterns",
            "budget_constraints": "Budget Constraints",
            "external_dependencies": "External Dependencies"
        }

        counts = {}
        for table, display_name in entity_tables.items():
            try:
                cursor.execute(f"SELECT COUNT(*) FROM {table}")
                counts[display_name] = cursor.fetchone()[0]
            except:
                counts[display_name] = 0

        return counts

    def _get_company_breakdown(self) -> Dict[str, Dict[str, int]]:
        """Get entity breakdown by company"""
        cursor = self.db.conn.cursor()

        # Get list of companies
        cursor.execute("SELECT DISTINCT company FROM interviews")
        companies = [row[0] for row in cursor.fetchall()]

        breakdown = {}
        for company in companies:
            breakdown[company] = {
                "interviews": 0,
                "pain_points": 0,
                "processes": 0,
                "systems": 0,
                "automation_candidates": 0
            }

            # Count interviews
            cursor.execute("SELECT COUNT(*) FROM interviews WHERE company = ?", (company,))
            breakdown[company]["interviews"] = cursor.fetchone()[0]

            # Count entities
            cursor.execute("SELECT COUNT(*) FROM pain_points WHERE company = ?", (company,))
            breakdown[company]["pain_points"] = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM processes WHERE company = ?", (company,))
            breakdown[company]["processes"] = cursor.fetchone()[0]

            try:
                cursor.execute("SELECT COUNT(*) FROM systems WHERE companies_using LIKE ?", (f"%{company}%",))
                breakdown[company]["systems"] = cursor.fetchone()[0]
            except:
                breakdown[company]["systems"] = 0

            cursor.execute("SELECT COUNT(*) FROM automation_candidates WHERE company = ?", (company,))
            breakdown[company]["automation_candidates"] = cursor.fetchone()[0]

        return breakdown

    def _get_quality_metrics(self) -> Dict[str, Any]:
        """Get data quality metrics"""
        cursor = self.db.conn.cursor()

        metrics = {
            "empty_descriptions": 0,
            "short_descriptions": 0,
            "encoding_issues": 0,
            "orphaned_entities": 0
        }

        # Check empty/short descriptions in pain_points
        cursor.execute("""
            SELECT COUNT(*) FROM pain_points
            WHERE description IS NULL OR description = ''
        """)
        metrics["empty_descriptions"] = cursor.fetchone()[0]

        cursor.execute("""
            SELECT COUNT(*) FROM pain_points
            WHERE LENGTH(description) < 20 AND description IS NOT NULL AND description != ''
        """)
        metrics["short_descriptions"] = cursor.fetchone()[0]

        # Check encoding issues
        cursor.execute("""
            SELECT COUNT(*) FROM pain_points
            WHERE description LIKE '%Ã¡%' OR description LIKE '%Ã©%'
        """)
        metrics["encoding_issues"] = cursor.fetchone()[0]

        # Check orphaned entities
        cursor.execute("""
            SELECT COUNT(*) FROM pain_points
            WHERE interview_id NOT IN (SELECT id FROM interviews)
        """)
        metrics["orphaned_entities"] = cursor.fetchone()[0]

        return metrics

    def _get_top_entities(self) -> Dict[str, List[Dict[str, Any]]]:
        """Get top entities by various criteria"""
        cursor = self.db.conn.cursor()

        top_entities = {}

        # Top pain points by severity
        cursor.execute("""
            SELECT type, description, severity, company
            FROM pain_points
            WHERE severity IN ('high', 'critical')
            ORDER BY
                CASE severity
                    WHEN 'critical' THEN 1
                    WHEN 'high' THEN 2
                    ELSE 3
                END
            LIMIT 10
        """)
        top_entities["critical_pain_points"] = [
            {
                "type": row[0],
                "description": row[1][:100],
                "severity": row[2],
                "company": row[3]
            }
            for row in cursor.fetchall()
        ]

        # Top automation candidates by impact
        cursor.execute("""
            SELECT name, process, impact, complexity, company
            FROM automation_candidates
            WHERE impact IN ('high', 'very high')
            ORDER BY
                CASE impact
                    WHEN 'very high' THEN 1
                    WHEN 'high' THEN 2
                    ELSE 3
                END
            LIMIT 10
        """)
        top_entities["high_impact_automations"] = [
            {
                "name": row[0],
                "process": row[1],
                "impact": row[2],
                "complexity": row[3],
                "company": row[4]
            }
            for row in cursor.fetchall()
        ]

        # Most used systems
        cursor.execute("""
            SELECT name, type, usage_count
            FROM systems
            ORDER BY usage_count DESC
            LIMIT 10
        """)
        top_entities["most_used_systems"] = [
            {
                "name": row[0],
                "type": row[1],
                "usage_count": row[2]
            }
            for row in cursor.fetchall()
        ]

        return top_entities

    def _get_extraction_status(self) -> Dict[str, int]:
        """Get extraction status breakdown"""
        cursor = self.db.conn.cursor()

        cursor.execute("""
            SELECT extraction_status, COUNT(*)
            FROM interviews
            GROUP BY extraction_status
        """)

        return dict(cursor.fetchall())

    def print_report(self):
        """Print human-readable report to console"""
        if not self.report_data:
            self.generate_full_report()

        data = self.report_data

        print(f"\n{'='*70}")
        print(f"📊 EXTRACTION REPORT")
        print(f"{'='*70}")
        print(f"Generated: {data['metadata']['generated_at']}")
        print(f"Database: {data['metadata']['database']}")

        print(f"\n📈 SUMMARY")
        print(f"{'-'*70}")
        summary = data['summary']
        print(f"Total Interviews: {summary['total_interviews']}")
        print(f"Total Companies: {summary['total_companies']}")
        print(f"Total Entities: {summary['total_entities']}")
        print(f"Avg Entities/Interview: {summary['avg_entities_per_interview']:.1f}")

        print(f"\n📋 ENTITY COUNTS")
        print(f"{'-'*70}")
        for entity_type, count in sorted(data['entity_counts'].items(), key=lambda x: x[1], reverse=True):
            print(f"{entity_type:30s}: {count:4d}")

        print(f"\n🏢 COMPANY BREAKDOWN")
        print(f"{'-'*70}")
        for company, stats in data['company_breakdown'].items():
            print(f"\n{company}:")
            print(f"  Interviews: {stats['interviews']}")
            print(f"  Pain Points: {stats['pain_points']}")
            print(f"  Processes: {stats['processes']}")
            print(f"  Systems: {stats['systems']}")
            print(f"  Automation Candidates: {stats['automation_candidates']}")

        print(f"\n✅ QUALITY METRICS")
        print(f"{'-'*70}")
        quality = data['quality_metrics']
        print(f"Empty Descriptions: {quality['empty_descriptions']}")
        print(f"Short Descriptions: {quality['short_descriptions']}")
        print(f"Encoding Issues: {quality['encoding_issues']}")
        print(f"Orphaned Entities: {quality['orphaned_entities']}")

        print(f"\n🔥 TOP CRITICAL PAIN POINTS")
        print(f"{'-'*70}")
        for i, pp in enumerate(data['top_entities']['critical_pain_points'][:5], 1):
            print(f"{i}. [{pp['severity'].upper()}] {pp['company']}")
            print(f"   {pp['description']}...")
            print()

        print(f"\n⚡ TOP AUTOMATION OPPORTUNITIES")
        print(f"{'-'*70}")
        for i, auto in enumerate(data['top_entities']['high_impact_automations'][:5], 1):
            print(f"{i}. {auto['name']} - {auto['company']}")
            print(f"   Impact: {auto['impact']} | Complexity: {auto['complexity']}")
            print()

        print(f"\n{'='*70}\n")

    def export_json(self, output_path: Path):
        """Export report to JSON file"""
        if not self.report_data:
            self.generate_full_report()

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(self.report_data, f, indent=2, ensure_ascii=False)

        print(f"✓ Report exported to: {output_path}")

    def export_excel(self, output_path: Path):
        """Export report to Excel file with multiple sheets"""
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            print("⚠️  Excel export requires: pip install pandas openpyxl")
            print("   Exporting to JSON instead...")
            json_path = output_path.with_suffix('.json')
            self.export_json(json_path)
            return

        if not self.report_data:
            self.generate_full_report()

        data = self.report_data

        # Create Excel writer
        output_path.parent.mkdir(parents=True, exist_ok=True)
        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        # Sheet 1: Summary
        summary_df = pd.DataFrame([data['summary']])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 2: Entity Counts
        entity_counts_df = pd.DataFrame(list(data['entity_counts'].items()), columns=['Entity Type', 'Count'])
        entity_counts_df = entity_counts_df.sort_values('Count', ascending=False)
        entity_counts_df.to_excel(writer, sheet_name='Entity Counts', index=False)

        # Sheet 3: Company Breakdown
        company_data = []
        for company, stats in data['company_breakdown'].items():
            row = {'Company': company}
            row.update(stats)
            company_data.append(row)
        company_df = pd.DataFrame(company_data)
        company_df.to_excel(writer, sheet_name='Company Breakdown', index=False)

        # Sheet 4: Quality Metrics
        quality_df = pd.DataFrame([data['quality_metrics']])
        quality_df.to_excel(writer, sheet_name='Quality Metrics', index=False)

        # Sheet 5: Top Pain Points
        if data['top_entities']['critical_pain_points']:
            pain_points_df = pd.DataFrame(data['top_entities']['critical_pain_points'])
            pain_points_df.to_excel(writer, sheet_name='Critical Pain Points', index=False)

        # Sheet 6: Top Automations
        if data['top_entities']['high_impact_automations']:
            automations_df = pd.DataFrame(data['top_entities']['high_impact_automations'])
            automations_df.to_excel(writer, sheet_name='Top Automations', index=False)

        # Save workbook
        writer.close()

        # Add formatting
        wb = load_workbook(output_path)

        # Format Summary sheet
        ws = wb['Summary']
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, color="FFFFFF")

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)
        print(f"✓ Excel report exported to: {output_path}")

    def close(self):
        """Close database connection"""
        self.db.close()


def main():
    """Main entry point"""
    import argparse

    parser = argparse.ArgumentParser(description="Generate comprehensive extraction report")
    parser.add_argument("--db", type=Path, default=DB_PATH, help="Path to database")
    parser.add_argument("--output", type=Path, help="Output file path (JSON or XLSX)")
    parser.add_argument("--format", choices=["json", "excel", "both"], default="both", help="Output format")
    parser.add_argument("--print", action="store_true", help="Print report to console")
    args = parser.parse_args()

    generator = ExtractionReportGenerator(args.db)

    try:
        # Generate report
        generator.generate_full_report()

        # Print to console
        if args.print or not args.output:
            generator.print_report()

        # Export to file
        if args.output:
            if args.format in ["json", "both"]:
                json_path = args.output.with_suffix('.json')
                generator.export_json(json_path)

            if args.format in ["excel", "both"]:
                excel_path = args.output.with_suffix('.xlsx')
                generator.export_excel(excel_path)
        else:
            # Default output
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_path = Path("reports") / f"extraction_report_{timestamp}"

            generator.export_json(default_path.with_suffix('.json'))
            generator.export_excel(default_path.with_suffix('.xlsx'))

    finally:
        generator.close()


if __name__ == "__main__":
    main()
